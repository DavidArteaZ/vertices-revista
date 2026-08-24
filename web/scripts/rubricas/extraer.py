"""Extrae los ocho instrumentos de dictamen del libro de Excel a extraidas.json.

    python3 web/scripts/rubricas/extraer.py

El libro (Vertices_BaseDatos_Editorial.xlsx) NO está en el repositorio: trae
nombres, correos y enlaces a manuscritos de autores reales. Lo que sí se
versiona es su salida, que son sólo etiquetas de rúbrica.

Qué se lee de cada hoja, y por qué de ahí:

  · fila 7  → las etiquetas. El prefijo ★ marca puerta eliminatoria o
              dimensión crítica; es la convención del propio libro.
  · las validaciones de datos → qué columnas son puertas ("Sí,No"), cuáles
              dimensiones ("0,1,2,3") y cuál admite N/A ("0,1,2,3,N/A").
              Se leen de ahí y no de la posición de la columna porque la
              frontera entre puertas y dimensiones se mueve de hoja en hoja.
  · A1      → el nivel (A, B o C).
  · S8/T8/U8/V8/W8 → las fórmulas de puntaje, máximo, puertas ★ OK,
              críticos ★ OK y la cascada de decisión. Se guardan tal cual
              para poder cotejarlas a mano contra lo que hace el motor.

Los umbrales de banda NO se extraen: viven transcritos en BANDAS, dentro de
generar-semillas.mjs, porque están enterrados en los IF anidados de W8 y
parsear esa fórmula sería más frágil que copiarla y comprobarla contra el
texto de la fila A2.
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parents[3]
LIBRO = RAIZ / "Vertices_BaseDatos_Editorial.xlsx"
SALIDA = Path(__file__).resolve().parent / "extraidas.json"


def columnas(rango: str) -> list[str]:
    """'J8:M57' -> ['J', 'K', 'L', 'M']"""
    a, b = rango.split(":")
    ca = re.match(r"([A-Z]+)", a).group(1)
    cb = re.match(r"([A-Z]+)", b).group(1)
    return [chr(x) for x in range(ord(ca), ord(cb) + 1)]


def main() -> int:
    if not LIBRO.exists():
        print(f"no encuentro {LIBRO}", file=sys.stderr)
        print("el libro no se versiona: contiene datos personales de autores", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(LIBRO)
    hojas = [s for s in wb.sheetnames if re.match(r"^\d\. ", s)]
    salida = []

    for h in hojas:
        ws = wb[h]
        cab = {
            re.match(r"([A-Z]+)", c.coordinate).group(1): c.value
            for c in ws[7]
            if c.value is not None
        }
        val = {str(dv.sqref): dv.formula1 for dv in ws.data_validations.dataValidation}

        puertas_cols: list[str] = []
        dim_cols: list[str] = []
        na_cols: list[str] = []
        for rango, f in val.items():
            if f == '"Sí,No"':
                puertas_cols += columnas(rango)
            elif f == '"0,1,2,3"':
                dim_cols += columnas(rango)
            elif f == '"0,1,2,3,N/A"':
                na_cols += columnas(rango)
        dim_cols = sorted(set(dim_cols + na_cols))

        salida.append({
            "hoja": h,
            "titulo": ws["A1"].value,
            "nivel": re.search(r"Nivel ([ABC])", ws["A1"].value).group(1),
            "bandas_texto": ws["A2"].value,
            "puertas": [
                {"col": c, "etiqueta": cab[c], "eliminatoria": cab[c].startswith("★")}
                for c in sorted(set(puertas_cols))
            ],
            "dimensiones": [
                {
                    "col": c,
                    "etiqueta": cab[c],
                    "critica": cab[c].startswith("★"),
                    "permite_na": c in na_cols,
                    "peso": 2 if "×2" in cab[c] else 1,
                }
                for c in dim_cols
            ],
            "etiquetas_banda": next(
                (f.strip('"').split(",") for r, f in val.items() if r.startswith("X")), []
            ),
            "formulas": {k: ws[k].value for k in ("S8", "T8", "U8", "V8", "W8")},
        })

    SALIDA.write_text(json.dumps(salida, ensure_ascii=False, indent=1) + "\n")
    print(f"escrito {SALIDA.relative_to(RAIZ)} — {len(salida)} instrumentos")
    for s in salida:
        estrella_p = sum(1 for p in s["puertas"] if p["eliminatoria"])
        estrella_d = sum(1 for d in s["dimensiones"] if d["critica"])
        print(
            f"  {s['hoja']:<28} nivel {s['nivel']}  "
            f"{len(s['puertas'])} puertas ({estrella_p} ★)  "
            f"{len(s['dimensiones'])} dims ({estrella_d} ★)"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
